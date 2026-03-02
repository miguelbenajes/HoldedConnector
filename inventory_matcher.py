#!/usr/bin/env python3
"""
inventory_matcher.py
---------------------
Fuzzy-matches invoice line items (sin product_id) against inventory products.
Generates an interactive Excel file for review/editing.

Usage:
    /usr/bin/python3 inventory_matcher.py

Output:
    products_to_import.xlsx (in current directory)
"""

import os
import sys
import difflib
import csv
from datetime import datetime

# Load .env
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    with open(_env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())

import connector  # noqa: E402

# Try to import openpyxl; if not available, fall back to CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("⚠️  openpyxl not found — falling back to CSV output")


def generate_excel_output(matched_rows, not_matched_rows):
    """Generate interactive Excel file with two sheets."""
    if not HAS_OPENPYXL:
        return None

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    matched_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    not_matched_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ─── Sheet 1: MATCHED ────────────────────────────────────────
    if matched_rows:
        ws_matched = wb.create_sheet("MATCHED")
        headers = [
            "✓ IMPORTED?",
            "Invoice Concept",
            "Times Billed",
            "Total Revenue",
            "Avg Price",
            "Matched Product",
            "Similarity %",
            "SKU Suggestion",
            "Price",
            "Stock",
            "Description"
        ]

        ws_matched.append(headers)

        # Format header
        for cell in ws_matched[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Add data rows
        for row_data in matched_rows:
            ws_matched.append(row_data)

        # Format data rows
        for idx, row in enumerate(ws_matched.iter_rows(min_row=2, max_row=len(matched_rows)+1), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                if col_idx == 1:  # Checkbox column
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.value = ""  # User will mark with X or ☑
                elif col_idx in [3, 4, 5, 7]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = matched_fill

        # Set column widths
        ws_matched.column_dimensions['A'].width = 14  # Checkbox
        ws_matched.column_dimensions['B'].width = 35  # Invoice concept
        ws_matched.column_dimensions['C'].width = 12  # Times
        ws_matched.column_dimensions['D'].width = 14  # Revenue
        ws_matched.column_dimensions['E'].width = 11  # Avg price
        ws_matched.column_dimensions['F'].width = 30  # Matched product
        ws_matched.column_dimensions['G'].width = 12  # Similarity
        ws_matched.column_dimensions['H'].width = 18  # SKU
        ws_matched.column_dimensions['I'].width = 11  # Price
        ws_matched.column_dimensions['J'].width = 10  # Stock
        ws_matched.column_dimensions['K'].width = 30  # Description

        # Freeze header row
        ws_matched.freeze_panes = "A2"

    # ─── Sheet 2: NOT_MATCHED ────────────────────────────────────
    if not_matched_rows:
        ws_not_matched = wb.create_sheet("NOT_MATCHED")
        headers = [
            "✓ CREATE?",
            "Invoice Concept",
            "Times Billed",
            "Total Revenue",
            "Avg Price",
            "SKU Suggestion",
            "Price",
            "Stock",
            "Description"
        ]

        ws_not_matched.append(headers)

        # Format header
        for cell in ws_not_matched[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Add data rows
        for row_data in not_matched_rows:
            ws_not_matched.append(row_data)

        # Format data rows
        for idx, row in enumerate(ws_not_matched.iter_rows(min_row=2, max_row=len(not_matched_rows)+1), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = border
                if col_idx == 1:  # Checkbox column
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.value = ""  # User will mark with X or ☑
                elif col_idx in [3, 4, 5]:  # Numeric columns
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = not_matched_fill

        # Set column widths
        ws_not_matched.column_dimensions['A'].width = 14  # Checkbox
        ws_not_matched.column_dimensions['B'].width = 35  # Invoice concept
        ws_not_matched.column_dimensions['C'].width = 12  # Times
        ws_not_matched.column_dimensions['D'].width = 14  # Revenue
        ws_not_matched.column_dimensions['E'].width = 11  # Avg price
        ws_not_matched.column_dimensions['F'].width = 18  # SKU
        ws_not_matched.column_dimensions['G'].width = 11  # Price
        ws_not_matched.column_dimensions['H'].width = 10  # Stock
        ws_not_matched.column_dimensions['I'].width = 30  # Description

        # Freeze header row
        ws_not_matched.freeze_panes = "A2"

    output_file = os.path.join(os.path.dirname(__file__), "products_to_import.xlsx")
    wb.save(output_file)
    return output_file


def generate_csv_output(matched_rows, not_matched_rows):
    """Fallback: generate CSV file."""
    output_file = os.path.join(os.path.dirname(__file__), "products_to_import.csv")

    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # Write MATCHED section
        if matched_rows:
            writer.writerow(["=== MATCHED PRODUCTS ==="])
            headers = [
                "Invoice Concept", "Times Billed", "Total Revenue", "Avg Price",
                "Matched Product", "Similarity %", "SKU Suggestion", "Price", "Stock", "Description"
            ]
            writer.writerow(headers)
            for row in matched_rows:
                writer.writerow(row[1:])  # Skip checkbox column
            writer.writerow([])

        # Write NOT_MATCHED section
        if not_matched_rows:
            writer.writerow(["=== NOT MATCHED (TO CREATE) ==="])
            headers = [
                "Invoice Concept", "Times Billed", "Total Revenue", "Avg Price",
                "SKU Suggestion", "Price", "Stock", "Description"
            ]
            writer.writerow(headers)
            for row in not_matched_rows:
                writer.writerow(row[1:])  # Skip checkbox column

    return output_file


def main():
    print("Connecting to database...")
    conn = connector.get_db()
    cur = connector._cursor(conn)

    # ─── Read all products ───────────────────────────────────────
    print("Reading products from inventory...")
    cur.execute("SELECT id, name, price, sku FROM products ORDER BY name")
    products = cur.fetchall()
    products_dict = {}
    prod_names_lower = []

    for p in products:
        if hasattr(p, 'keys'):
            prod_id = p['id']
            prod_name = p['name']
            prod_price = float(p['price'] or 0)
            prod_sku = p['sku']
        else:
            prod_id, prod_name, prod_price, prod_sku = p

        products_dict[prod_name.lower()] = {
            'id': prod_id,
            'name': prod_name,
            'price': prod_price,
            'sku': prod_sku
        }
        prod_names_lower.append(prod_name.lower())

    print(f"  Loaded {len(products_dict)} products")

    # ─── Read invoice items without product_id ───────────────────
    print("Reading invoice concepts without product_id...")
    cur.execute("""
        SELECT
            name,
            COUNT(*) as veces,
            COALESCE(SUM(subtotal), 0) as revenue_total,
            COALESCE(AVG(price), 0) as precio_medio
        FROM invoice_items
        WHERE product_id IS NULL
          AND name IS NOT NULL
          AND TRIM(name) != ''
        GROUP BY name
        ORDER BY revenue_total DESC
    """)

    invoice_items = cur.fetchall()
    conn.close()

    print(f"  Found {len(invoice_items)} unique invoice concepts")

    # ─── Fuzzy match ────────────────────────────────────────────
    print("Fuzzy matching...")
    matched_rows = []
    not_matched_rows = []

    for item in invoice_items:
        if hasattr(item, 'keys'):
            concept = item['name']
            veces = int(item['veces'])
            revenue = float(item['revenue_total'])
            precio = float(item['precio_medio'])
        else:
            concept, veces, revenue, precio = item

        # Fuzzy match
        matches = difflib.get_close_matches(concept.lower(), prod_names_lower, n=1, cutoff=0.60)

        # Generate SKU suggestion
        sku_suggest = concept.strip().upper().replace(' ', '_')[:20]

        if matches:
            # MATCHED
            idx = prod_names_lower.index(matches[0])
            matched_prod = list(products_dict.values())[idx]
            ratio = difflib.SequenceMatcher(None, concept.lower(), matches[0]).ratio()
            sim_pct = int(ratio * 100)

            row = [
                "",  # Checkbox (empty, user fills)
                concept,
                veces,
                f"{revenue:.2f}",
                f"{precio:.2f}",
                matched_prod['name'],
                f"{sim_pct}%",
                matched_prod['sku'] or sku_suggest,
                matched_prod['price'],
                "",  # Stock
                ""  # Description
            ]
            matched_rows.append(row)
        else:
            # NOT MATCHED
            row = [
                "",  # Checkbox (empty, user fills)
                concept,
                veces,
                f"{revenue:.2f}",
                f"{precio:.2f}",
                sku_suggest,
                f"{precio:.2f}" if precio > 0 else "",  # Suggest invoice avg price
                "",  # Stock
                ""  # Description
            ]
            not_matched_rows.append(row)

    # ─── Generate output ────────────────────────────────────────
    if HAS_OPENPYXL:
        print("Generating Excel file...")
        output_file = generate_excel_output(matched_rows, not_matched_rows)
        print(f"\n✓ Generated: {output_file}")
    else:
        print("Generating CSV file...")
        output_file = generate_csv_output(matched_rows, not_matched_rows)
        print(f"\n✓ Generated: {output_file}")

    # Summary
    print("\n" + "=" * 60)
    print(f"Matched products:     {len(matched_rows)}")
    print(f"Not matched (new):    {len(not_matched_rows)}")
    print(f"Total invoice items:  {len(matched_rows) + len(not_matched_rows)}")

    total_revenue = sum(float(row[3]) for row in matched_rows + not_matched_rows)
    print(f"Total revenue:        €{total_revenue:.2f}")
    print("=" * 60)
    print("\nNext steps:")
    print("1. Open the Excel file")
    print("2. Review MATCHED products (validate similarity %)")
    print("3. Go to NOT_MATCHED sheet")
    print("4. Edit SKU, Price, Stock, Description as needed")
    print("5. Mark with ✓ or X in the first column for products you want to create")
    print("6. Import to Holded UI or dashboard")


if __name__ == "__main__":
    main()
