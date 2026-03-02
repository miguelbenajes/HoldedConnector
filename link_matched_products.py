#!/usr/bin/env python3
"""
link_matched_products.py
------------------------
Reads products_to_import.xlsx (MATCHED sheet) and:
1. Updates invoice_items to link concepts without product_id to actual products
2. Creates amortizations for each matched product
3. All within a single transaction for atomicity

Usage:
    /usr/bin/python3 link_matched_products.py

Requirements:
    - products_to_import.xlsx must exist in current directory
    - openpyxl must be installed (for Excel reading)
    - Supabase DATABASE_URL must be set in .env
"""

import os
import sys
from datetime import datetime

# Load .env
_env_path = os.path.join(os.path.dirname(__file__), ".env")
if os.path.exists(_env_path):
    with open(_env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip())

import connector  # noqa: E402

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed")
    sys.exit(1)


def main():
    excel_path = os.path.join(os.path.dirname(__file__), "products_to_import.xlsx")

    if not os.path.exists(excel_path):
        print(f"ERROR: {excel_path} not found")
        sys.exit(1)

    print("Reading Excel file...")
    wb = load_workbook(excel_path)

    if "MATCHED" not in wb.sheetnames:
        print("ERROR: 'MATCHED' sheet not found in Excel")
        sys.exit(1)

    ws = wb["MATCHED"]

    # Parse header
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    print(f"Found headers: {list(headers.keys())}")

    # Read data rows
    matched_items = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        invoice_concept = row[headers["Invoice Concept"] - 1].value
        matched_product = row[headers["Matched Product"] - 1].value

        if not invoice_concept or not matched_product:
            continue

        matched_items.append({
            "invoice_concept": str(invoice_concept).strip(),
            "matched_product": str(matched_product).strip(),
        })

    print(f"Loaded {len(matched_items)} matched items from Excel")

    if not matched_items:
        print("No items to process")
        return

    # ─── Connect to Supabase ────────────────────────────────────────────────
    print("\nConnecting to Supabase...")
    conn = connector.get_db()

    try:
        cur = connector._cursor(conn)

        # ─── Build mappings ─────────────────────────────────────────────────
        print("Building product name → ID mapping...")
        cur.execute("SELECT id, name FROM products")
        prod_rows = cur.fetchall()
        prod_map = {}  # name (lower) → id

        for p in prod_rows:
            if hasattr(p, 'keys'):
                prod_id = p['id']
                prod_name = p['name']
            else:
                prod_id, prod_name = p

            prod_map[prod_name.lower()] = prod_id

        print(f"  Loaded {len(prod_map)} products")

        # ─── Process matches ────────────────────────────────────────────────
        print("\nProcessing matches...")

        updated_items = 0
        created_amortizations = 0
        errors = []
        processed_products = set()  # Track unique products for amortization

        for item in matched_items:
            invoice_concept = item["invoice_concept"]
            matched_product = item["matched_product"]

            # Find product_id
            prod_id = prod_map.get(matched_product.lower())
            if not prod_id:
                errors.append(f"Product not found: {matched_product}")
                continue

            # UPDATE invoice_items: link concept to product
            try:
                cur.execute(
                    connector._q(
                        "UPDATE invoice_items SET product_id = ? "
                        "WHERE name = ? AND product_id IS NULL"
                    ),
                    (prod_id, invoice_concept),
                )
                rows_updated = cur.rowcount
                if rows_updated > 0:
                    updated_items += rows_updated
                    print(f"  ✓ Linked {rows_updated} invoice items: '{invoice_concept}' → {matched_product}")
            except Exception as e:
                errors.append(f"Failed to update invoice_items for '{invoice_concept}': {e}")
                continue

            # Track product for amortization creation
            if prod_id not in processed_products:
                processed_products.add(prod_id)

        # ─── Create amortizations for processed products ──────────────────
        print("\nCreating amortizations...")

        for prod_id in processed_products:
            # Find product details
            prod_details = None
            for p in prod_rows:
                if hasattr(p, 'keys'):
                    if p['id'] == prod_id:
                        prod_details = {'name': p['name']}
                        break
                else:
                    if p[0] == prod_id:
                        prod_details = {'name': p[1]}
                        break

            if not prod_details:
                continue

            # INSERT amortization (ON CONFLICT handles duplicates)
            try:
                cur.execute(
                    connector._q(
                        "INSERT INTO amortizations "
                        "(product_id, product_name, purchase_price, purchase_date, notes, product_type, created_at) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?) "
                        "ON CONFLICT (product_id) DO NOTHING"
                    ),
                    (
                        prod_id,
                        prod_details['name'],
                        0,  # purchase_price = 0 (no cost known yet)
                        datetime.now().isoformat(),
                        "Auto-linked from invoice concepts",
                        "alquiler",
                        datetime.now().isoformat(),
                    ),
                )
                rows_inserted = cur.rowcount
                if rows_inserted > 0:
                    created_amortizations += 1
                    print(f"  ✓ Created amortization: {prod_details['name']}")
                else:
                    print(f"  ~ Amortization already exists: {prod_details['name']}")
            except Exception as e:
                errors.append(f"Failed to create amortization for {prod_id}: {e}")
                continue

        # ─── Commit ─────────────────────────────────────────────────────────
        print("\nCommitting transaction...")
        conn.commit()

        # ─── Summary ─────────────────────────────────────────────────────────
        print("\n" + "=" * 70)
        print("OPERATION COMPLETED")
        print("=" * 70)
        print(f"✓ Invoice items updated: {updated_items}")
        print(f"✓ Amortizations created: {created_amortizations}")

        if errors:
            print(f"\n⚠️  Errors encountered ({len(errors)}):")
            for error in errors:
                print(f"  - {error}")

        # ─── Verify ─────────────────────────────────────────────────────────
        print("\nVerifying...")
        cur.execute(connector._q("SELECT COUNT(*) as c FROM amortizations"))
        total_amort = connector._fetch_one_val(cur, 'c')
        cur.execute(
            connector._q(
                "SELECT COUNT(*) as c FROM invoice_items "
                "WHERE product_id IS NOT NULL"
            )
        )
        linked_items = connector._fetch_one_val(cur, 'c')

        print(f"Total amortizations in database: {total_amort}")
        print(f"Total invoice items with product_id: {linked_items}")

        print("\n" + "=" * 70)
        print("Next steps:")
        print("1. Open http://localhost:8000/amortizations in dashboard")
        print("2. Verify the new products appear with revenue calculated")
        print("3. Check GET /api/amortizations/summary for total stats")
        print("=" * 70)

    except Exception as e:
        print(f"\nERROR: {e}")
        conn.rollback()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
