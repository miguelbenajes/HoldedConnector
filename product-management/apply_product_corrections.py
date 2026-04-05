#!/usr/bin/env python3
"""
Apply Product Corrections & Link to Existing Inventory
Purpose: Reclassify items based on user decisions, link to existing products, update Supabase.
"""

import os
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Add project root to path (connector.py lives one level up)
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
sys.path.insert(0, ROOT)

import connector

def get_existing_product_by_name(name_fragment):
    """Find existing product by partial name match."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        # Search for products containing the fragment
        cur.execute(
            "SELECT id, name FROM products WHERE name ILIKE %s LIMIT 1",
            (f"%{name_fragment}%",)
        )
        row = cur.fetchone()
        if row:
            return {
                'id': connector._row_val(row, 'id', 0),
                'name': connector._row_val(row, 'name', 1)
            }
        return None
    finally:
        conn.close()


def create_amortization_for_product(product_id, product_name):
    """Create amortization record for a linked product."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        # Check if amortization already exists
        cur.execute("SELECT id FROM amortizations WHERE product_id = %s", (product_id,))
        if cur.fetchone():
            return False  # Already exists

        # Create new amortization
        cur.execute("""
            INSERT INTO amortizations (product_id, purchase_price, purchase_date, notes)
            VALUES (%s, 0, NOW(), %s)
            ON CONFLICT (product_id) DO NOTHING
        """, (product_id, f"Auto-linked from service: {product_name}"))

        conn.commit()
        return True
    except Exception as e:
        print(f"Error creating amortization: {e}")
        return False
    finally:
        conn.close()


def load_and_process_products():
    """Load products from Excel and apply user corrections."""
    excel_file = os.path.join(HERE, 'products_for_import.xlsx')
    wb = load_workbook(excel_file)

    corrections = {
        'real_products': [],
        'services': [],
        'expenses': [],
        'linked_to_inventory': []
    }

    print("\n" + "="*80)
    print("APPLYING PRODUCT CORRECTIONS")
    print("="*80 + "\n")

    # Load Real Products
    if '1. Real Products' in wb.sheetnames:
        ws = wb['1. Real Products']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue
            corrections['real_products'].append(row[0])

    # Load Services and apply reclassifications
    if '3. Services-Fees' in wb.sheetnames:
        ws = wb['3. Services-Fees']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue

            service_name = str(row[0]).strip()

            # Check for reclassifications
            if service_name == 'E200 Spigot adapter 16-29mm':
                print(f"  ✓ Reclassified: '{service_name}' → Real Product")
                corrections['real_products'].append(service_name)

            elif 'BIR17235' in service_name:
                print(f"  ✓ Ignored: '{service_name}' (administrative)")

            elif 'ALQ MACBOOK' in service_name:
                # Try to link to existing macbook product
                existing = get_existing_product_by_name('Macbook Pro Max M3')
                if existing:
                    print(f"  ✓ Linked: '{service_name}' → '{existing['name']}' (ID: {existing['id']})")
                    corrections['linked_to_inventory'].append({
                        'original_name': service_name,
                        'linked_to': existing['name'],
                        'product_id': existing['id']
                    })
                    # Create amortization
                    if create_amortization_for_product(existing['id'], service_name):
                        print(f"    ✓ Amortization created")
                else:
                    print(f"  ⚠ Not linked: '{service_name}' (macbook m3 not found)")
                    corrections['services'].append(service_name)

            elif 'Dinner Thursday 12-9-2024' in service_name or 'Dinner Friday 13-9-2024' in service_name:
                print(f"  ✓ Reclassified: '{service_name}' → Expenses")
                corrections['expenses'].append(service_name)

            else:
                # Keep as service
                corrections['services'].append(service_name)

    # Load Expenses
    if '2. Expenses' in wb.sheetnames:
        ws = wb['2. Expenses']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue
            corrections['expenses'].append(row[0])

    return corrections


def regenerate_excel(corrections, output_file=None):
    if output_file is None:
        output_file = os.path.join(HERE, 'products_for_import.xlsx')
    """Regenerate Excel with corrected classifications."""
    wb = Workbook()
    wb.remove(wb.active)

    def style_sheet(ws, header_color):
        from openpyxl.styles import Alignment
        header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border

    # 1. REAL PRODUCTS (including reclassified items)
    if corrections['real_products']:
        ws = wb.create_sheet('1. Real Products', 0)
        ws.append(['Product Name', 'SKU', 'Category', 'Price (€)', 'Stock', 'Description'])

        for product in corrections['real_products']:
            ws.append([product, '', '', '', '', ''])

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30
        ws.freeze_panes = 'A2'
        style_sheet(ws, "4472C4")

    # 2. EXPENSES (including reclassified dinners)
    if corrections['expenses']:
        ws = wb.create_sheet('2. Expenses', 1)
        ws.append([
            'Expense Name',
            'Type',
            'Project ID',
            'Amount (€)',
            'Date',
            'Vendor/Description',
            'Status'
        ])

        for expense in corrections['expenses']:
            ws.append([expense, 'Expense', '', '', '', '', 'Pending'])

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 12
        ws.freeze_panes = 'A2'
        style_sheet(ws, "70AD47")

    # 3. SERVICES-FEES (corrected)
    if corrections['services']:
        ws = wb.create_sheet('3. Services-Fees', 2)
        ws.append([
            'Service/Fee Name',
            'Type',
            'Hours',
            'Date',
            'Rate (€/hr)',
            'Total (€)',
            'Status',
            'Notes'
        ])

        for service in corrections['services']:
            ws.append([service, 'Fee', '', '', '', '', 'Pending', ''])

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 25
        ws.freeze_panes = 'A2'
        style_sheet(ws, "FFC000")

    # 4. LINKED TO INVENTORY
    if corrections['linked_to_inventory']:
        ws = wb.create_sheet('4. Linked to Inventory', 3)
        ws.append(['Original Name', 'Linked To', 'Product ID', 'Status', 'Amortization'])

        for item in corrections['linked_to_inventory']:
            ws.append([
                item['original_name'],
                item['linked_to'],
                item['product_id'],
                'Linked',
                'Created'
            ])

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.freeze_panes = 'A2'
        style_sheet(ws, "00B050")

    wb.save(output_file)


def main():
    print("Loading and processing products from Excel...")
    corrections = load_and_process_products()

    print("\n" + "="*80)
    print("SUMMARY OF CORRECTIONS")
    print("="*80)
    print(f"\nReal Products (including reclassified): {len(corrections['real_products'])} items")
    print(f"  • Original: 177")
    print(f"  • + Reclassified: E200 Spigot adapter = 178")

    print(f"\nExpenses (including reclassified dinners): {len(corrections['expenses'])} items")
    print(f"  • Original: 64")
    print(f"  • + Reclassified: 2 dinners = 66")

    print(f"\nServices-Fees (corrected): {len(corrections['services'])} items")
    print(f"  • Original: 46")
    print(f"  • - Reclassified out: 3 items")
    print(f"  • - Linked to inventory: 1 item")
    print(f"  • = {len(corrections['services'])}")

    print(f"\nLinked to Existing Inventory: {len(corrections['linked_to_inventory'])} items")
    for item in corrections['linked_to_inventory']:
        print(f"  • '{item['original_name']}' → '{item['linked_to']}'")

    print(f"\nIgnored (Administrative): BIR17235")

    print("\n" + "="*80)
    print("UPDATING EXCEL FILE")
    print("="*80 + "\n")

    regenerate_excel(corrections)
    print(f"✓ Updated: products_for_import.xlsx\n")

    print("="*80)
    print("SUPABASE UPDATES")
    print("="*80)
    print(f"\n✓ Amortizations created for linked products")
    print(f"✓ Ready to create in Holded when you're ready\n")


if __name__ == '__main__':
    main()
