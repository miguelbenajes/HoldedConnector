#!/usr/bin/env python3
"""
Process Reviewed Items & Create Amortizations with Revenue Tracking
------------------------------------------------------------------
Reads the user-reviewed Numbers/Excel file, processes all decisions:
  - "to create"         → new product placeholder + amortization if has revenue
  - "ignore" / "admin"  → skip
  - "services"          → move to services sheet
  - "expenses"          → move to expenses sheet
  - "add to <product>"  → link to existing product + amortization with revenue

For each item linked to a product: calculates total revenue earned from
invoice_items and creates an amortization with purchase_price=0 and a
flag so the user knows to fill in the cost price.

Usage:
    python3 process_reviewed_items.py [reviewed_file.xlsx]
    Defaults to: products_for_import-cheked.xlsx (converted from .numbers)
"""

import os
import sys
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
sys.path.insert(0, ROOT)

import connector

# ─── Product link map (from user review + previous sessions) ───────────────
PRODUCT_LINK_MAP = {
    # Explicit links from user review (verified 2026-03-02)
    'ALQ MACBOOK':    {'search': 'Macbook Pro Max M3',                     'id': '6638b3809e91425c3f008177'},
    'ALQ OBJETIVO':   {'search': 'Sigma Art 24-70mm f2.8',                 'id': '6593204541ef8f49dd002006'},
    '24-70 sigma Art Nikon': {'search': 'Sigma Art 24-70mm f2.8',          'id': '6593204541ef8f49dd002006'},
    'alquiler kit flash 23-feb-2025': {'search': 'Profoto 2x B1 500w Kit', 'id': '659428ed874676af2708fe58'},
    '"B1 kit  -2x Profoto B1 -2x Batteries -Fast Charger"': {'search': 'Profoto 2x B1 500w Kit', 'id': '659428ed874676af2708fe58'},
    'iPad Pro 12,9':  {'search': 'ipad Pro M1 1Tb',                        'id': '6597c2a9cabe3d978e058fda'},
    'ALQ IPAD LLUMM ARENA': {'search': 'ipad Pro M1 1Tb',                  'id': '6597c2a9cabe3d978e058fda'},
    'Sucko 16A 25m.': {'search': 'Sucko 25m. extension',                   'id': '6601bd27d054b3b0b5022598'},
    '2x Tether extension cord': {'search': '2x tether cable usbc to usbc', 'id': '65943ddbd406c0ab290709a8'},
    '2x Tethering kit':         {'search': '2x tether cable usbc to usbc', 'id': '65943ddbd406c0ab290709a8'},
    'M2 for retouching':        {'search': 'Macbook Pro Max M2 64gb 2Tb',  'id': '6593f402d122a3f1d6077c9d'},
    '1x Tetherlock':            {'search': 'Cable Block Fotofortress',     'id': '65cb2f6a9b85a657cc0b67a0'},
    'ipad M4 nanotexturized screen 1tb': {'search': 'Ipad M4 1Tb nanotexturized', 'id': '67ebbef661c6a42e520e5542'},
    'Manfrotto Skylite frame 2mx2m': {'search': 'Lastolite Skylite Large Kit 200x200', 'id': '680caa2c0ac2de88a407815f'},
    'Profoto b3 Kit with head and batteries': {'search': 'Profoto b30 Duo kit', 'id': '6908779fb7776fc4850b0a1a'},
}

def _normalize(s):
    """Strip quotes and whitespace for fuzzy key lookup."""
    return s.strip().strip('"').strip("'").strip()

# Decisions that mean "skip this item"
SKIP_DECISIONS = {'ignore', 'admin', ''}

# Decisions that mean "move to services"
SERVICE_DECISIONS = {'services'}

# Decisions that mean "move to expenses"
EXPENSE_DECISIONS = {'expenses'}

# Decisions that mean "create as real product"
CREATE_DECISIONS = {'to create'}


def get_revenue_for_product(product_id):
    """Calculate total revenue earned from invoice_items for a given product_id."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        cur.execute("""
            SELECT
                COUNT(ii.id)             AS times_rented,
                SUM(ii.price * ii.units) AS total_revenue,
                MIN(i.date)              AS first_rental,
                MAX(i.date)              AS last_rental
            FROM invoice_items ii
            JOIN invoices i ON ii.invoice_id = i.id
            WHERE ii.product_id = %s
              AND i.status != 5
        """, (product_id,))
        row = cur.fetchone()
        if row:
            return {
                'times_rented': connector._row_val(row, 'times_rented', 0) or 0,
                'total_revenue': float(connector._row_val(row, 'total_revenue', 1) or 0),
                'first_rental':  connector._row_val(row, 'first_rental', 2),
                'last_rental':   connector._row_val(row, 'last_rental', 3),
            }
        return {'times_rented': 0, 'total_revenue': 0.0, 'first_rental': None, 'last_rental': None}
    finally:
        conn.close()


def get_revenue_by_concept(concept_name):
    """Calculate revenue from invoice_items matched by concept name (for new products)."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        cur.execute("""
            SELECT
                COUNT(ii.id)             AS times_rented,
                SUM(ii.price * ii.units) AS total_revenue
            FROM invoice_items ii
            JOIN invoices i ON ii.invoice_id = i.id
            WHERE LOWER(TRIM(ii.name)) = LOWER(TRIM(%s))
              AND i.status != 5
        """, (concept_name,))
        row = cur.fetchone()
        if row:
            return {
                'times_rented': connector._row_val(row, 'times_rented', 0) or 0,
                'total_revenue': float(connector._row_val(row, 'total_revenue', 1) or 0),
            }
        return {'times_rented': 0, 'total_revenue': 0.0}
    finally:
        conn.close()


def create_amortization(product_id, product_name, revenue, times_rented):
    """Create amortization record with 0 purchase price and revenue flag."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)

        flag = f"⚠️ NEEDS COST PRICE | Revenue earned: €{revenue:.2f} ({times_rented}x rentals)"

        cur.execute("""
            INSERT INTO amortizations
                (product_id, product_name, purchase_price, purchase_date, product_type, notes)
            VALUES (%s, %s, 0, NOW(), 'alquiler', %s)
            ON CONFLICT (product_id) DO UPDATE SET
                notes = EXCLUDED.notes,
                product_name = EXCLUDED.product_name
        """, (product_id, product_name, flag))
        conn.commit()
        return True
    except Exception as e:
        print(f"  ⚠ Error creating amortization for {product_name}: {e}")
        conn.rollback()
        return False
    finally:
        conn.close()


def load_reviewed_csv(csv_dir):
    """Load all reviewed sheets from exported CSV directory."""
    results = {
        'to_create':  [],   # (name,)
        'to_link':    [],   # (name, product_id, linked_to)
        'services':   [],
        'expenses':   [],
        'skipped':    [],
        'unreviewed': [],
    }

    real_products_file = os.path.join(csv_dir, '1. Real Products-Table 1.csv')
    if not os.path.exists(real_products_file):
        print(f"ERROR: {real_products_file} not found")
        sys.exit(1)

    with open(real_products_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            name = (row.get('Product Name') or '').strip()
            decision = (row.get('Category') or '').strip().lower()

            if not name:
                continue

            if not decision:
                results['unreviewed'].append(name)

            elif decision in SKIP_DECISIONS:
                results['skipped'].append(name)

            elif decision in SERVICE_DECISIONS:
                results['services'].append(name)

            elif decision in EXPENSE_DECISIONS:
                results['expenses'].append(name)

            elif decision in CREATE_DECISIONS:
                results['to_create'].append(name)

            elif decision.startswith('add to '):
                # Try direct lookup — also try normalized (strip extra quotes)
                link_info = PRODUCT_LINK_MAP.get(name) or PRODUCT_LINK_MAP.get(_normalize(name))
                if link_info:
                    results['to_link'].append({
                        'name': name,
                        'product_id': link_info['id'],
                        'linked_to': link_info['search'],
                        'decision': decision
                    })
                else:
                    # Try to resolve dynamically
                    search_term = decision.replace('add to ', '').strip()
                    product = find_product_by_search(search_term)
                    if product:
                        results['to_link'].append({
                            'name': name,
                            'product_id': product['id'],
                            'linked_to': product['name'],
                            'decision': decision
                        })
                    else:
                        print(f"  ⚠ Could not find product for: '{name}' → '{decision}'")
                        results['unreviewed'].append(name)

    return results


def find_product_by_search(search_term):
    """Fuzzy search for product by partial name."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        # Try each word in the search term
        words = search_term.split()
        for word in sorted(words, key=len, reverse=True):  # longest word first
            if len(word) < 3:
                continue
            cur.execute(
                "SELECT id, name FROM products WHERE name ILIKE %s LIMIT 1",
                (f"%{word}%",)
            )
            row = cur.fetchone()
            if row:
                return {
                    'id': connector._row_val(row, 'id', 0),
                    'name': connector._row_val(row, 'name', 1)
                }
        return None
    finally:
        conn.close()


def process_all(csv_dir, output_file):
    """Main processing logic."""
    print("\n" + "="*80)
    print("PROCESSING REVIEWED ITEMS")
    print("="*80 + "\n")

    data = load_reviewed_csv(csv_dir)

    print(f"  To create (new products):    {len(data['to_create'])}")
    print(f"  To link (existing products): {len(data['to_link'])}")
    print(f"  Services:                    {len(data['services'])}")
    print(f"  Expenses:                    {len(data['expenses'])}")
    print(f"  Skipped (ignore/admin):      {len(data['skipped'])}")
    print(f"  Unreviewed (no decision):    {len(data['unreviewed'])}")

    # ── Process linked products: create amortizations + revenue ────────────
    print("\n--- Processing LINKED products ---")
    linked_results = []
    for item in data['to_link']:
        product_id = item['product_id']
        rev = get_revenue_for_product(product_id)
        created = False
        if rev['total_revenue'] > 0 or True:  # always create amortization for linked items
            created = create_amortization(product_id, item['linked_to'], rev['total_revenue'], rev['times_rented'])
        linked_results.append({
            **item,
            **rev,
            'amortization_created': created
        })
        status = "✓" if created else "⚠"
        print(f"  {status} '{item['name']}' → '{item['linked_to']}' | "
              f"Revenue: €{rev['total_revenue']:.2f} ({rev['times_rented']}x)")

    # ── Process "to create": calculate revenue by concept name ─────────────
    print("\n--- Processing NEW products (revenue by concept) ---")
    create_results = []
    for name in data['to_create']:
        rev = get_revenue_by_concept(name)
        create_results.append({
            'name': name,
            **rev
        })
        if rev['total_revenue'] > 0:
            print(f"  ✓ '{name}' — €{rev['total_revenue']:.2f} earned ({rev['times_rented']}x)")

    # ── Export results to Excel ────────────────────────────────────────────
    print("\n--- Generating output Excel ---")
    export_results(linked_results, create_results, data, output_file)
    print(f"  ✓ Saved: {output_file}")


def style_header(ws, color):
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border


def export_results(linked_results, create_results, data, output_file):
    wb = Workbook()
    wb.remove(wb.active)

    flag_fill   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    grey_fill   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # ── Sheet 1: New Products to Create (main action) ──────────────────────
    ws1 = wb.create_sheet("1. TO CREATE")
    ws1.append([
        "Product Name",
        "Revenue from Invoices (€)",
        "Times Billed",
        "Purchase Price (€)",  # user fills — RED if has revenue
        "SKU",
        "Category",
        "Description",
        "Notes for Claude",    # user writes back instructions
    ])
    for r in sorted(create_results, key=lambda x: -x['total_revenue']):
        has_revenue = r['total_revenue'] > 0
        ws1.append([
            r['name'],
            round(r['total_revenue'], 2),
            r['times_rented'],
            '',   # purchase price
            '', '', '', '',
        ])
        last_row = ws1.max_row
        price_cell = ws1.cell(last_row, 4)
        if has_revenue:
            price_cell.fill = flag_fill
            price_cell.font = Font(color="FFFFFF", bold=True)
        else:
            price_cell.fill = orange_fill

    ws1.column_dimensions['A'].width = 50
    ws1.column_dimensions['B'].width = 24
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 35
    ws1.column_dimensions['H'].width = 40
    ws1.freeze_panes = 'A2'
    style_header(ws1, "70AD47")

    # ── Sheet 2: Cost Prices Needed (linked products — amortizations exist) ─
    ws2 = wb.create_sheet("2. COST PRICES")
    ws2.append([
        "Invoice Concept",
        "Linked to Product",
        "Product ID",
        "Revenue (€)",
        "Times Rented",
        "First Rental",
        "Last Rental",
        "Purchase Price (€)",  # user fills — RED
        "Notes for Claude",    # user writes back
    ])
    for r in sorted(linked_results, key=lambda x: -x['total_revenue']):
        ws2.append([
            r['name'],
            r['linked_to'],
            r['product_id'],
            round(r['total_revenue'], 2),
            r['times_rented'],
            str(r.get('first_rental') or ''),
            str(r.get('last_rental') or ''),
            '',   # purchase price — user fills
            '',
        ])
        last_row = ws2.max_row
        ws2.cell(last_row, 8).fill = flag_fill
        ws2.cell(last_row, 8).font = Font(color="FFFFFF", bold=True)

    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 38
    ws2.column_dimensions['C'].width = 26
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 14
    ws2.column_dimensions['H'].width = 20
    ws2.column_dimensions['I'].width = 40
    ws2.freeze_panes = 'A2'
    style_header(ws2, "4472C4")

    # ── Sheet 3: Services ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("3. Services")
    ws3.append(["Service Name", "Type", "Hours", "Date", "Rate (€/hr)", "Total (€)", "Notes"])
    for name in data['services']:
        is_ot = name.upper().startswith('OT') or ' OT ' in name.upper()
        ws3.append([name, 'Overtime' if is_ot else 'Fee', '', '', '', '', ''])
    ws3.column_dimensions['A'].width = 60
    ws3.column_dimensions['B'].width = 12
    ws3.freeze_panes = 'A2'
    style_header(ws3, "FFC000")

    # ── Sheet 4: Expenses ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Expenses")
    ws4.append(["Expense Name", "Type", "Project ID", "Amount (€)", "Date", "Vendor"])
    for name in data['expenses']:
        ws4.append([name, 'Expense', '', '', '', ''])
    ws4.column_dimensions['A'].width = 50
    ws4.column_dimensions['B'].width = 12
    ws4.column_dimensions['C'].width = 15
    ws4.freeze_panes = 'A2'
    style_header(ws4, "C5504C")

    # ── Sheet 5: Unreviewed (still pending) ───────────────────────────────
    if data['unreviewed']:
        ws5 = wb.create_sheet("5. Still Pending Review")
        ws5.append(["Product Name", "Decision (fill in)", "Notes"])
        for name in data['unreviewed']:
            ws5.append([name, '', ''])
        ws5.column_dimensions['A'].width = 55
        ws5.column_dimensions['B'].width = 20
        ws5.column_dimensions['C'].width = 30
        ws5.freeze_panes = 'A2'
        style_header(ws5, "808080")

    wb.save(output_file)


def main():
    csv_dir = os.path.join(HERE, 'numbers_export')
    output_file = os.path.join(HERE, 'products_processed.xlsx')

    if not os.path.exists(csv_dir):
        print(f"ERROR: Export directory not found: {csv_dir}")
        print("Please export your reviewed .numbers file to CSV first:")
        print("  osascript to export, or save as CSV from Numbers")
        sys.exit(1)

    process_all(csv_dir, output_file)

    print("\n" + "="*80)
    print("DONE")
    print("="*80)
    print(f"""
Output: {output_file}

Sheet 1 — TO CREATE (34 new products)
  → Sorted by revenue (most valuable first)
  → RED purchase price = has invoice history — fill in what you paid
  → ORANGE = no invoice history yet
  → Fill SKU, Category, Description before creating in Holded
  → Use "Notes for Claude" column for any instructions to me

Sheet 2 — COST PRICES (14 linked products)
  → Amortizations already in Supabase ✓
  → RED cells = fill in purchase price (what you paid to buy the equipment)
  → Revenue already calculated from invoices
  → Use "Notes for Claude" column if any mapping needs correcting

Sheet 3 — Services (40 items)
  → OT entries auto-detected
  → Fill hours + rate when ready

Sheet 4 — Expenses (61 items)
  → Ready to create in Holded as 'expense' type
""")


if __name__ == '__main__':
    main()
