#!/usr/bin/env python3
"""
classify_products.py
--------------------
Analyzes 307 NOT_MATCHED products from products_to_import.xlsx
and classifies them into categories:
- Real Products (equipment/inventory to create)
- Services (labor, consulting)
- Expenses (transport, food, misc)
- Administrative (notes, invalid entries)

Output: Excel file with classifications and recommendations

Usage:
    /usr/bin/python3 classify_products.py
"""

import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# products_to_import.xlsx lives in the project root (one level up)
ROOT_DIR = os.path.join(os.path.dirname(__file__), "..")
EXCEL_PATH = os.path.join(ROOT_DIR, "products_to_import.xlsx")

# Classification keywords
SERVICE_KEYWORDS = [
    "fee", "assist", "fee", "ot ", "day ", "session", "project", "consulting",
    "hora", "hora", "hour", "labor", "jornada", "shooting", "book"
]

EXPENSE_KEYWORDS = [
    "taxi", "parking", "petrol", "fuel", "courier", "shipping", "cena", "dinner",
    "diner", "glovo", "meal", "transport", "delivery", "invoice", "ticket",
    "train", "flight", "transavia", "hotel", "accommodation"
]

ADMIN_KEYWORDS = [
    "anulacion", "annulation", "project code", "transavia", "ref",
    "invoice", "reference"
]

# Real product keywords (positive indicators)
PRODUCT_KEYWORDS = [
    "camera", "lens", "battery", "cable", "filter", "reflector", "light",
    "stand", "tripod", "clamp", "mount", "softbox", "diffusion", "umbrella",
    "kit", "pro", "kit", "ssd", "drive", "keyboard", "mouse", "hub",
    "monitor", "screen", "ipad", "macbook", "phone", "tablet", "charger",
    "adapter", "converter", "controller", "drone", "gimbal", "stabilizer"
]


def classify_product(concept):
    """Classify a product concept into a category."""
    concept_lower = concept.lower()

    # Check admin first (most specific)
    if any(kw in concept_lower for kw in ADMIN_KEYWORDS):
        return "Administrative", "⚪"

    # Check services
    if any(kw in concept_lower for kw in SERVICE_KEYWORDS):
        return "Service", "🔴"

    # Check expenses
    if any(kw in concept_lower for kw in EXPENSE_KEYWORDS):
        return "Expense", "🟡"

    # Check for real products
    product_score = sum(1 for kw in PRODUCT_KEYWORDS if kw in concept_lower)

    if product_score >= 1:
        return "Real Product", "🟢"

    # Default: Real Product (assume it's inventory unless proven otherwise)
    return "Real Product", "🟢"


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: {EXCEL_PATH} not found")
        return

    print(f"Reading {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH)
    ws_not_matched = wb["NOT_MATCHED"]

    # Parse NOT_MATCHED sheet
    products = []
    for row_idx, row in enumerate(ws_not_matched.iter_rows(min_row=2, values_only=False), start=2):
        concept = row[1].value if len(row) > 1 else None
        times = row[2].value if len(row) > 2 else 0
        revenue = row[3].value if len(row) > 3 else 0
        avg_price = row[4].value if len(row) > 4 else 0
        sku = row[5].value if len(row) > 5 else ""
        price = row[6].value if len(row) > 6 else ""
        stock = row[7].value if len(row) > 7 else ""
        desc = row[8].value if len(row) > 8 else ""

        if not concept:
            continue

        category, emoji = classify_product(str(concept))

        products.append({
            'concept': str(concept).strip(),
            'times': times or 0,
            'revenue': float(str(revenue).replace('€', '').strip()) if revenue else 0.0,
            'avg_price': float(str(avg_price).replace('€', '').strip()) if avg_price else 0.0,
            'sku': str(sku).strip() if sku else '',
            'price': str(price).strip() if price else '',
            'stock': str(stock).strip() if stock else '',
            'desc': str(desc).strip() if desc else '',
            'category': category,
            'emoji': emoji,
        })

    print(f"Classified {len(products)} products\n")

    # Generate summary
    categories = {}
    for p in products:
        cat = p['category']
        if cat not in categories:
            categories[cat] = {'count': 0, 'revenue': 0.0, 'products': []}
        categories[cat]['count'] += 1
        categories[cat]['revenue'] += p['revenue']
        categories[cat]['products'].append(p)

    print("=" * 80)
    print("CLASSIFICATION SUMMARY")
    print("=" * 80)
    for cat in ['Real Product', 'Service', 'Expense', 'Administrative']:
        if cat in categories:
            stats = categories[cat]
            print(f"\n{cat}: {stats['count']} items | Revenue: €{stats['revenue']:.2f}")
            if cat == 'Real Product':
                print("  → These should be created in Holded inventory")
            elif cat == 'Service':
                print("  → These are labor/time (track separately, not in inventory)")
            elif cat == 'Expense':
                print("  → These are operating expenses (use as product type 'expense')")
            elif cat == 'Administrative':
                print("  → These are notes/invalid entries (skip these)")

    print("\n" + "=" * 80)

    # Create output Excel
    output_path = os.path.join(os.path.dirname(__file__), "products_classified.xlsx")
    output_wb = Workbook()
    output_wb.remove(output_wb.active)

    # Colors for categories
    colors = {
        'Real Product': 'C6EFCE',  # Light green
        'Service': 'FFC7CE',       # Light red
        'Expense': 'FFEB9C',       # Light yellow
        'Administrative': 'E2EFDA' # Very light
    }

    # Create sheet per category
    for category in ['Real Product', 'Service', 'Expense', 'Administrative']:
        if category not in categories:
            continue

        ws = output_wb.create_sheet(title=category[:20])

        # Headers
        headers = ['CONCEPT', 'TIMES', 'REVENUE', 'AVG PRICE', 'SKU', 'PRICE', 'STOCK', 'DESC']
        ws.append(headers)

        # Format header
        fill = PatternFill(start_color=colors[category], end_color=colors[category], fill_type='solid')
        font = Font(bold=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Add data rows
        for p in categories[category]['products']:
            ws.append([
                p['concept'],
                p['times'],
                f"€{p['revenue']:.2f}",
                f"€{p['avg_price']:.2f}",
                p['sku'],
                p['price'],
                p['stock'],
                p['desc']
            ])

        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=len(categories[category]['products'])+1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 30

    output_wb.save(output_path)
    print(f"\n✓ Saved: {output_path}")

    # Print recommendations
    print("\n" + "=" * 80)
    print("RECOMMENDATIONS")
    print("=" * 80)

    real_count = categories.get('Real Product', {}).get('count', 0)
    service_count = categories.get('Service', {}).get('count', 0)
    expense_count = categories.get('Expense', {}).get('count', 0)
    admin_count = categories.get('Administrative', {}).get('count', 0)

    print(f"\n✅ CREATE IN HOLDED: {real_count} products")
    print(f"   - These are inventory items (equipment, materials, etc.)")
    print(f"\n⏭️  SKIP / RECLASSIFY: {service_count} services")
    print(f"   - These are labor/time entries, not products")
    print(f"\n💰 CREATE AS EXPENSE TYPE: {expense_count} expenses")
    print(f"   - Operating costs, not inventory")
    print(f"\n🗑️  DELETE / IGNORE: {admin_count} administrative entries")
    print(f"   - Notes, references, invalid entries")
    print("\n" + "=" * 80)


if __name__ == "__main__":
    main()
