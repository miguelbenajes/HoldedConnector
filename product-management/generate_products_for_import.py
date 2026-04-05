#!/usr/bin/env python3
"""
Generate Master Excel for Product Import & Data Entry
Purpose: Create comprehensive Excel file for manual review and Holded import.
Includes: Real Products, Expenses (with project_id), Fees (with overtime detection), Administrative.
"""

import os
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Add project root to path (connector.py lives one level up)
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
sys.path.insert(0, ROOT)

import connector

def get_existing_products():
    """Fetch all products from database for reference."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        cur.execute("SELECT id, name FROM products ORDER BY name")
        products = []
        for row in cur.fetchall():
            product_id = connector._row_val(row, 'id', 0)
            product_name = connector._row_val(row, 'name', 1)
            products.append((product_id, product_name))
        return products
    finally:
        conn.close()


def detect_overtime(text):
    """Detect if text contains overtime notation (OT xx hours on xx day)."""
    pattern = r'OT\s+(\d+(?:\.\d+)?)\s*h(?:ours?)?\s+(?:on|en)\s+(.+)'
    match = re.search(pattern, str(text), re.IGNORECASE)
    return match is not None, match.groups() if match else (None, None)


def load_classified_products(filepath):
    """Load classified products from Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    data = {
        'expenses': [],
        'services': [],
        'real_products': [],
        'administrative': []
    }

    # Load Expense sheet
    if 'Expense' in wb.sheetnames:
        ws = wb['Expense']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue
            data['expenses'].append(row[0])

    # Load Service sheet
    if 'Service' in wb.sheetnames:
        ws = wb['Service']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue
            data['services'].append(row[0])

    # Load Real Product sheet
    if 'Real Product' in wb.sheetnames:
        ws = wb['Real Product']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue
            data['real_products'].append(row[0])

    # Load Administrative sheet
    if 'Administrative' in wb.sheetnames:
        ws = wb['Administrative']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1 or not row or not row[0]:
                continue
            data['administrative'].append(row[0])

    return data


def style_sheet(ws, header_color):
    """Apply consistent styling to a worksheet."""
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Apply border to all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_products_for_import(data, output_file):
    """Generate master Excel file for product import."""
    wb = Workbook()
    wb.remove(wb.active)

    # Get existing products for reference
    existing_products = get_existing_products()

    # ========== 1. REAL PRODUCTS (to create in Holded) ==========
    if data['real_products']:
        ws = wb.create_sheet('1. Real Products', 0)
        ws.append(['Product Name', 'SKU', 'Category', 'Price (€)', 'Stock', 'Description'])

        for product_name in data['real_products']:
            ws.append([product_name, '', '', '', '', ''])

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 30

        # Style
        style_sheet(ws, "4472C4")  # Blue

        # Freeze header
        ws.freeze_panes = 'A2'

    # ========== 2. EXPENSES (with project_id) ==========
    if data['expenses']:
        ws = wb.create_sheet('2. Expenses', 1)
        ws.append([
            'Expense Name',
            'Type',
            'Project ID',
            'Amount (€)',
            'Date',
            'Vendor/Description',
            'Status'
        ])

        for expense_name in data['expenses']:
            ws.append([expense_name, 'Expense', '', '', '', '', 'Pending'])

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 12

        # Style
        style_sheet(ws, "70AD47")  # Green

        # Freeze header
        ws.freeze_panes = 'A2'

        # Add note about project_id
        ws.append([''])
        ws.append(['Note: Fill in Project ID if this expense is associated with a specific project'])
        ws['A10'].font = Font(italic=True, size=9)

    # ========== 3. SERVICES / FEES (with overtime detection) ==========
    if data['services']:
        ws = wb.create_sheet('3. Services-Fees', 2)

        # Separate overtime from regular services
        overtime_services = []
        regular_services = []

        for service_name in data['services']:
            is_overtime, (hours, day) = detect_overtime(service_name)
            if is_overtime:
                overtime_services.append((service_name, hours, day))
            else:
                regular_services.append(service_name)

        # Header
        ws.append([
            'Service/Fee Name',
            'Type',
            'Hours',
            'Date',
            'Rate (€/hr)',
            'Total (€)',
            'Status',
            'Notes'
        ])

        # Regular services first
        for service_name in regular_services:
            ws.append([service_name, 'Fee', '', '', '', '', 'Pending', ''])

        # Overtime services (pre-filled with detected values)
        for service_name, hours, day in overtime_services:
            ws.append([
                service_name,
                'Overtime',
                hours if hours else '',
                day if day else '',
                '',
                '',
                'Pending',
                'Auto-detected OT entry'
            ])

        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 25

        # Style
        style_sheet(ws, "FFC000")  # Gold

        # Freeze header
        ws.freeze_panes = 'A2'

        # Add info
        ws.append([''])
        ws.append([f'Total Services: {len(data["services"])} | Overtime detected: {len(overtime_services)} | Regular: {len(regular_services)}'])
        ws['A10'].font = Font(italic=True, size=9)

    # ========== 4. ADMINISTRATIVE (for review) ==========
    if data['administrative']:
        ws = wb.create_sheet('4. Administrative', 3)
        ws.append(['Item', 'Original Text', 'Action', 'Reason', 'Status'])

        for admin_item in data['administrative']:
            ws.append([admin_item, admin_item, '', '', 'Pending Review'])

        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15

        # Style
        style_sheet(ws, "C5504C")  # Red

        # Freeze header
        ws.freeze_panes = 'A2'

    # ========== 5. REFERENCE - Existing Products ==========
    if existing_products:
        ws = wb.create_sheet('Reference - Existing', 4)
        ws.append(['Product ID', 'Product Name'])

        for product_id, product_name in sorted(existing_products, key=lambda x: x[1])[:100]:  # First 100
            ws.append([product_id, product_name])

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 50

        # Style
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

        ws.append([''])
        ws.append([f'Showing first 100 of {len(existing_products)} existing products'])

    wb.save(output_file)


def main():
    classified_file = os.path.join(HERE, 'products_classified.xlsx')
    output_file = os.path.join(HERE, 'products_for_import.xlsx')

    print("Loading classified products...")
    data = load_classified_products(classified_file)

    print("\n" + "="*80)
    print("GENERATING IMPORT FILE")
    print("="*80 + "\n")

    print(f"Real Products:      {len(data['real_products'])} items")
    print(f"Expenses:           {len(data['expenses'])} items")
    print(f"Services/Fees:      {len(data['services'])} items")

    # Detect overtime in services
    overtime_count = sum(1 for s in data['services'] if detect_overtime(s)[0])
    regular_count = len(data['services']) - overtime_count
    print(f"  ├─ Overtime entries detected: {overtime_count}")
    print(f"  └─ Regular services:          {regular_count}")

    print(f"Administrative:     {len(data['administrative'])} items")

    print("\n" + "="*80 + "\n")

    print(f"✓ Generating: {output_file}")
    export_products_for_import(data, output_file)
    print(f"✓ Done!\n")

    print("="*80)
    print("FILE STRUCTURE")
    print("="*80)
    print("""
Sheet 1: Real Products
  → Fill in: SKU, Category, Price, Stock, Description
  → Then create these in Holded

Sheet 2: Expenses
  → Fill in: Project ID (optional), Amount, Date, Vendor
  → Then create as 'expense' type in Holded

Sheet 3: Services/Fees
  → Overtime entries auto-detected (look for "OT xx hours on xx day")
  → Fill in: Hours, Date, Rate, Total
  → Then create as 'fee' type in Holded

Sheet 4: Administrative
  → Review each item
  → Mark decision: Create / Delete / Link to existing
  → Will ask what to do after you review

Sheet 5: Reference
  → Your existing 228 products (for reference)
  → Use to manually link items if needed

""" + "="*80 + "\n")


if __name__ == '__main__':
    main()
