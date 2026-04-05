#!/usr/bin/env python3
"""
Match Expenses to Existing Inventory Products
Purpose: Try to link unmatched expenses to existing products via fuzzy matching.
If no match found, export for manual review.
"""

import os
import sys
from difflib import get_close_matches
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Add project root to path (connector.py lives one level up)
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, "..")
sys.path.insert(0, ROOT)

import connector

def get_existing_products():
    """Fetch all products from database with name and ID."""
    conn = connector.get_db()
    try:
        cur = connector._cursor(conn)
        cur.execute("SELECT id, name FROM products ORDER BY name")
        products = {}
        for row in cur.fetchall():
            product_id = connector._row_val(row, 'id', 0)
            product_name = connector._row_val(row, 'name', 1)
            if product_name:
                products[product_name.strip()] = product_id
        return products
    finally:
        conn.close()


def load_classified_products(filepath):
    """Load classified products from Excel file."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath)

    # Get different sheets
    expenses = []
    real_products = []
    services = []

    # Load Expense sheet
    if 'Expense' in wb.sheetnames:
        ws = wb['Expense']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:  # Skip header
                continue
            if row and row[0]:  # Skip empty rows
                expenses.append(row[0])

    # Load Real Product sheet
    if 'Real Product' in wb.sheetnames:
        ws = wb['Real Product']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:  # Skip header
                continue
            if row and row[0]:
                real_products.append(row[0])

    # Load Service sheet
    if 'Service' in wb.sheetnames:
        ws = wb['Service']
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if idx == 1:  # Skip header
                continue
            if row and row[0]:
                services.append(row[0])

    return {
        'expenses': expenses,
        'real_products': real_products,
        'services': services
    }


def fuzzy_match_expenses(expenses, existing_products, threshold=0.65):
    """Match expenses to existing products. Return matched and unmatched."""
    matched = []
    unmatched = []

    for expense in expenses:
        expense_clean = str(expense).strip()

        # Try fuzzy matching
        close_matches = get_close_matches(
            expense_clean,
            list(existing_products.keys()),
            n=1,
            cutoff=threshold
        )

        if close_matches:
            matched_product = close_matches[0]
            product_id = existing_products[matched_product]
            matched.append({
                'original_name': expense_clean,
                'matched_product': matched_product,
                'product_id': product_id,
                'similarity': round(
                    len(set(expense_clean) & set(matched_product)) /
                    max(len(expense_clean), len(matched_product)), 2
                )
            })
        else:
            unmatched.append(expense_clean)

    return matched, unmatched


def export_results(matched, unmatched, real_products, services, output_file):
    """Export matched/unmatched expenses to Excel."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # 1. MATCHED EXPENSES (can link to existing products)
    if matched:
        ws = wb.create_sheet('Matched Expenses')
        ws.append(['Expense Name', 'Matched Product', 'Product ID', 'Similarity %'])

        for item in matched:
            ws.append([
                item['original_name'],
                item['matched_product'],
                item['product_id'],
                f"{int(item['similarity']*100)}%"
            ])

        # Style header
        header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # 2. UNMATCHED EXPENSES (needs manual review)
    if unmatched:
        ws = wb.create_sheet('Unmatched Expenses')
        ws.append(['Expense Name', 'Action', 'Notes'])

        for expense in unmatched:
            ws.append([expense, '', ''])

        # Style header
        header_fill = PatternFill(start_color="FFB6C6", end_color="FFB6C6", fill_type="solid")  # Light red
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # 3. REAL PRODUCTS (ready to create in Holded)
    if real_products:
        ws = wb.create_sheet('Real Products Ready')
        ws.append(['Product Name', 'Status'])

        for product in real_products:
            ws.append([product, 'Ready to create'])

        # Style header
        header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Light blue
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # 4. SERVICES (keep unmapped)
    if services:
        ws = wb.create_sheet('Services')
        ws.append(['Service Name', 'Status'])

        for service in services:
            ws.append([service, 'Keep unmapped'])

        # Style header
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_file)


def main():
    classified_file = os.path.join(HERE, 'products_classified.xlsx')
    output_file = os.path.join(HERE, 'products_final_review.xlsx')

    print("Loading classified products...")
    classified = load_classified_products(classified_file)

    print("Loading existing inventory products...")
    existing_products = get_existing_products()
    print(f"Found {len(existing_products)} existing products in inventory")

    print("\n" + "="*80)
    print("MATCHING EXPENSES TO EXISTING INVENTORY")
    print("="*80 + "\n")

    expenses = classified['expenses']
    print(f"Processing {len(expenses)} expenses...")

    matched, unmatched = fuzzy_match_expenses(expenses, existing_products, threshold=0.65)

    print(f"\n  ✅ Matched to existing products: {len(matched)}")
    if matched:
        for item in matched[:5]:  # Show first 5
            print(f"     • '{item['original_name']}' → '{item['matched_product']}' ({int(item['similarity']*100)}%)")
        if len(matched) > 5:
            print(f"     ... and {len(matched) - 5} more")

    print(f"\n  ❌ Unmatched (need manual review): {len(unmatched)}")
    if unmatched:
        for expense in unmatched[:5]:  # Show first 5
            print(f"     • '{expense}'")
        if len(unmatched) > 5:
            print(f"     ... and {len(unmatched) - 5} more")

    print("\n" + "="*80)
    print("SUMMARY")
    print("="*80)
    print(f"Real Products (ready to create):    {len(classified['real_products'])} items")
    print(f"Services (keep unmapped):           {len(classified['services'])} items")
    print(f"Expenses matched to inventory:      {len(matched)} items")
    print(f"Expenses needing manual review:     {len(unmatched)} items")
    print("\n" + "="*80 + "\n")

    # Export
    print(f"✓ Saving results to: {output_file}")
    export_results(
        matched,
        unmatched,
        classified['real_products'],
        classified['services'],
        output_file
    )
    print(f"✓ Done!\n")

    # Recommendations
    print("RECOMMENDATIONS")
    print("="*80)
    print(f"\n1. CREATE IN HOLDED: {len(classified['real_products'])} real products")
    print(f"   → Open 'Real Products Ready' sheet in {output_file}")
    print(f"\n2. LINK TO INVENTORY: {len(matched)} expenses")
    print(f"   → Open 'Matched Expenses' sheet in {output_file}")
    print(f"   → Verify the matches look correct")
    print(f"\n3. MANUAL REVIEW: {len(unmatched)} expenses")
    print(f"   → Open 'Unmatched Expenses' sheet in {output_file}")
    print(f"   → Decide if each should be:")
    print(f"      a) Created as new product")
    print(f"      b) Linked to a different existing product")
    print(f"      c) Marked as service/fee")
    print(f"      d) Deleted (not needed)")
    print(f"\n4. SERVICES: {len(classified['services'])} items to handle separately")
    print("   → Open 'Services' sheet in {output_file}")
    print("\n" + "="*80 + "\n")


if __name__ == '__main__':
    main()
